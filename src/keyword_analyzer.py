# src/keyword_analyzer.py (超高速化・RuntimeError修正版)

import pandas as pd
import tkinter as tk
from tkinter import ttk, font, scrolledtext
from tkinter import filedialog, messagebox
import os
from datetime import datetime
from src.serp_analyzer import SerpAnalyzer
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import concurrent.futures
import queue

class KeywordAnalyzer:
    def __init__(self, serp_analyzer: SerpAnalyzer):
        self.serp_analyzer = serp_analyzer
        self.dialog = None
        self.progressbar = None
        self.info_labels = None
        self.df_to_analyze = None
        self.analysis_results = []
        self.output_path = ""
        self.completed_count = 0
        self.results_queue = queue.Queue()
        print("[OK] KeywordAnalyzerの初期化に成功しました。（高速・安定版）")

    def _select_csv_file(self):
        root = tk.Tk()
        root.withdraw()
        initial_dir = r"C:\Users\ktmno\Desktop\auto blog\KWラッコエクセル"
        os.makedirs(initial_dir, exist_ok=True)
        return filedialog.askopenfilename(
            title="ラッコキーワードのCSVファイルを選択してください",
            filetypes=[("CSV files", "*.csv")],
            initialdir=initial_dir
        )

    def _load_and_prepare_data(self, file_path):
        if not file_path: return None
        df_raw = None
        try:
            # 3段階のエンコーディング試し
            try:
                # 1. UTF-8で試す (一般的なCSV)
                df_raw = pd.read_csv(file_path, encoding='utf-8')
            except UnicodeDecodeError:
                try:
                    # 2. Shift-JISで試す (Excel日本語版の古いCSV)
                    print("INFO: UTF-8での読み込みに失敗。Shift-JIS(cp932)で再試行します。")
                    df_raw = pd.read_csv(file_path, encoding='cp932')
                except UnicodeDecodeError:
                    # 3. UTF-16 (タブ区切り)で試す (ラッコキーワードのTSV)
                    print("INFO: Shift-JISでの読み込みに失敗。UTF-16(タブ区切り)で再試行します。")
                    df_raw = pd.read_csv(file_path, encoding='utf-16', sep='\t')
            
            if df_raw is None:
                raise ValueError("全てのエンコーディングでCSVの読み込みに失敗しました。")

            required_cols = ['キーワード', '月間検索数']
            if not all(col in df_raw.columns for col in required_cols):
                messagebox.showerror("エラー", "CSVファイルに必要な列（'キーワード', '月間検索数'）が見つかりません。")
                return None
            df = df_raw[required_cols].copy()
            df.columns = ['keyword', 'avg_monthly_searches']
            df.dropna(subset=['keyword'], inplace=True)
            df['keyword'] = df['keyword'].astype(str)
            df.drop_duplicates(subset=['keyword'], keep='first', inplace=True)
            df['avg_monthly_searches'] = df['avg_monthly_searches'].astype(str).str.split('-').str[0]
            df['avg_monthly_searches'] = pd.to_numeric(df['avg_monthly_searches'], errors='coerce')
            df.dropna(subset=['avg_monthly_searches'], inplace=True)
            df = df[df['avg_monthly_searches'] > 0]
            df['avg_monthly_searches'] = df['avg_monthly_searches'].astype(int)
            df_filtered = df[(df['avg_monthly_searches'] >= 500) & (df['avg_monthly_searches'] <= 3000)].copy()
            if df_filtered.empty:
                messagebox.showwarning("情報", "分析対象（月間検索ボリューム500～3000）のキーワードが見つかりませんでした。")
                return None
            print(f"[OK] {len(df_filtered)} 件の分析対象キーワードを抽出しました。（500～3000件/月）")
            return df_filtered
        except Exception as e:
            messagebox.showerror("エラー", f"CSV読み込みエラー:\n{e}")
            return None

    def _calculate_aim(self, allintitle, intitle):
        is_allintitle_valid = pd.notna(allintitle) and isinstance(allintitle, (int, float))
        is_intitle_valid = pd.notna(intitle) and isinstance(intitle, (int, float))
        if is_allintitle_valid and is_intitle_valid and allintitle <= 10 and intitle <= 30000:
            return '[OK]'
        return '[NG]'

    def _create_progress_dialog(self, total):
        self.dialog = tk.Toplevel()
        self.dialog.title("競合分析中...")
        self.dialog.geometry("600x350")
        self.dialog.resizable(False, False)
        self.dialog.attributes("-topmost", True)
        main_font = font.Font(family="Meiryo UI", size=10)
        bold_font = font.Font(family="Meiryo UI", size=11, weight="bold")
        style = ttk.Style(self.dialog)
        style.layout('text.Horizontal.TProgressbar',
                     [('Horizontal.Progressbar.trough',
                       {'children': [('Horizontal.Progressbar.pbar',
                                     {'side': 'left', 'sticky': 'ns'})],
                        'sticky': 'nswe'}),
                      ('Horizontal.Progressbar.label', {'sticky': ''})])
        style.configure('text.Horizontal.TProgressbar', text=' 0 %', anchor='center', font=main_font)
        ttk.Label(self.dialog, text="分析の進捗:", font=main_font).pack(pady=(10,0))
        self.progressbar = ttk.Progressbar(self.dialog, maximum=total, length=550, style='text.Horizontal.TProgressbar')
        self.progressbar.pack(pady=5, padx=20)
        ttk.Separator(self.dialog, orient='horizontal').pack(fill='x', pady=10, padx=20)
        info_frame = ttk.Frame(self.dialog)
        info_frame.pack(pady=5, padx=20, fill='x')
        self.info_labels = {}
        label_texts = ["現在のキーワード:", "allintitle:", "intitle:", "Q&Aサイト:", "SNS:", "無料ブログ:"]
        for i, text in enumerate(label_texts):
            ttk.Label(info_frame, text=text, font=bold_font).grid(row=i, column=0, sticky='w', padx=5, pady=2)
            self.info_labels[text] = ttk.Label(info_frame, text="-", font=main_font, anchor='w')
            self.info_labels[text].grid(row=i, column=1, sticky='w', padx=5, pady=2)

    def _create_keyword_selection_dialog(self, keywords: list[str]):
        self.selected_keywords = None
        self.selection_dialog_is_submitted = False
        dialog = tk.Toplevel()
        dialog.title("分析対象キーワードの編集・選択")
        dialog.geometry("700x600")
        dialog.attributes("-topmost", True)
        main_frame = tk.Frame(dialog, padx=15, pady=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        label_text = """CSVから抽出された分析候補キーワードです。
不要な行を削除したり、内容を編集してから分析を開始してください。"""
        tk.Label(main_frame, text=label_text, anchor='w', justify='left').pack(fill=tk.X, pady=(0, 10))
        text_widget = scrolledtext.ScrolledText(main_frame, wrap=tk.WORD, height=25)
        text_widget.pack(fill=tk.BOTH, expand=True)
        text_widget.insert(tk.END, "\n".join(keywords))
        text_widget.focus_set()
        def on_confirm():
            edited_text = text_widget.get("1.0", tk.END).strip()
            self.selected_keywords = [line.strip() for line in edited_text.splitlines() if line.strip()]
            self.selection_dialog_is_submitted = True
            dialog.destroy()
        def on_cancel():
            self.selection_dialog_is_submitted = False
            dialog.destroy()
        button_frame = tk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(15, 0))
        confirm_button = tk.Button(button_frame, text="この内容で分析開始", command=on_confirm, height=2)
        confirm_button.pack(fill=tk.X)
        dialog.protocol("WM_DELETE_WINDOW", on_cancel)
        dialog.grab_set()
        dialog.wait_window()
        return self.selected_keywords if self.selection_dialog_is_submitted else None

    def _format_excel_output(self, filepath):
        workbook = load_workbook(filepath)
        worksheet = workbook.active
        header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        aim_ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        aim_ng_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        rival_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            worksheet.column_dimensions[cell.column_letter].width = 18
        worksheet.column_dimensions['A'].width = 35
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.column == 8:
                    cell.alignment = center_align
                    if cell.value == '[OK]': cell.fill = aim_ok_fill
                    else: cell.fill = aim_ng_fill
                if cell.column in [5, 6, 7] and cell.value is not None and str(cell.value).isdigit():
                    cell.fill = rival_fill
        workbook.save(filepath)

    def _analyze_keyword_concurrently(self, row):
        """ワーカースレッドで実行されるキーワード分析処理"""
        keyword = row['keyword']
        allintitle, intitle, weak_ranks = self.serp_analyzer.analyze_top10_serps(keyword)
        aim_judgement = self._calculate_aim(allintitle, intitle)
        result_row = {
            'keyword': keyword, 'avg_monthly_searches': row['avg_monthly_searches'],
            'allintitle': allintitle, 'intitle': intitle,
            'Q&Aサイト': weak_ranks.get('Q&Aサイト'), 'SNS': weak_ranks.get('SNS'),
            '無料ブログ': weak_ranks.get('無料ブログ'), 'AIM判定': aim_judgement
        }
        self.results_queue.put(result_row)

    def _finalize_and_save(self):
        """分析結果をExcelに保存する最終処理"""
        print("\nすべての分析が完了しました。結果を処理しています...")
        df_final = pd.DataFrame(self.analysis_results)
        df_final.sort_values(
            by=['AIM判定', 'Q&Aサイト', 'SNS', '無料ブログ'],
            ascending=[False, True, True, True],
            na_position='last',
            inplace=True
        )
        final_columns = ['keyword', 'avg_monthly_searches', 'allintitle', 'intitle', 'Q&Aサイト', 'SNS', '無料ブログ', 'AIM判定']
        df_final = df_final[final_columns]
        try:
            df_final.to_excel(self.output_path, index=False, engine='openpyxl')
            self._format_excel_output(self.output_path)
            success_message = f"[DONE] 分析完了！\n結果を装飾・ソートしてExcelファイルに出力しました:\n{self.output_path}"
            print(f"\n{success_message}")
            messagebox.showinfo("完了", success_message)
        except Exception as e:
            messagebox.showerror("エラー", f"Excelファイルへの書き込み中にエラーが発生しました。\n\n詳細: {e}")

    def _process_queue(self):
        """キューを監視し、GUIを安全に更新するポーリングメソッド"""
        try:
            result = self.results_queue.get_nowait()
            self.completed_count += 1
            self.analysis_results.append(result)
            num_to_analyze = len(self.df_to_analyze)
            self.progressbar['value'] = self.completed_count
            percentage = (self.completed_count / num_to_analyze) * 100
            style = ttk.Style(self.dialog)
            style.configure('text.Horizontal.TProgressbar', text=f'{self.completed_count} / {num_to_analyze} ({percentage:.0f} %)')
            self.info_labels["現在のキーワード:"].config(text=result['keyword'])
            self.info_labels["allintitle:"].config(text=str(result['allintitle']) if pd.notna(result['allintitle']) else "-")
            self.info_labels["intitle:"].config(text=str(result['intitle']) if pd.notna(result['intitle']) else "-")
            self.info_labels["Q&Aサイト:"].config(text=str(result.get('Q&Aサイト', '-')))
            self.info_labels["SNS:"].config(text=str(result.get('SNS', '-')))
            self.info_labels["無料ブログ:"].config(text=str(result.get('無料ブログ', '-')))
        except queue.Empty:
            pass
        
        if self.completed_count < len(self.df_to_analyze):
            self.dialog.after(100, self._process_queue)
        else:
            # GUIのメインループを終了させるために、dialog.quit()を呼ぶ
            self.dialog.quit()

    def run_analysis(self):
        csv_path = self._select_csv_file()
        if not csv_path: return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(csv_path))[0]
        output_dir = os.path.dirname(csv_path)
        self.output_path = os.path.join(output_dir, f"{base_name}_{timestamp}_競合分析済み.xlsx")
        
        df_filtered = self._load_and_prepare_data(csv_path)
        if df_filtered is None or df_filtered.empty: return
        
        initial_keywords = df_filtered['keyword'].tolist()
        selected_keywords = self._create_keyword_selection_dialog(initial_keywords)
        if selected_keywords is None:
            print("[NG] キーワード選択がキャンセルされました。処理を中断します。")
            return
            
        self.df_to_analyze = df_filtered[df_filtered['keyword'].isin(selected_keywords)].copy()
        self.df_to_analyze['keyword'] = pd.Categorical(self.df_to_analyze['keyword'], categories=selected_keywords, ordered=True)
        self.df_to_analyze.sort_values('keyword', inplace=True)
        self.df_to_analyze.reset_index(drop=True, inplace=True)
        
        if self.df_to_analyze.empty:
            print("ℹ️ 分析対象のキーワードがありません。処理を終了します。")
            return
            
        num_to_run = len(self.df_to_analyze)
        print(f"\n[OK] {num_to_run} 件のキーワードを超高速で分析します。")
        
        self.analysis_results = []
        self.completed_count = 0
        while not self.results_queue.empty():
            self.results_queue.get()

        self._create_progress_dialog(num_to_run)

        # ワーカースレッドプールを開始
        executor = concurrent.futures.ThreadPoolExecutor(max_workers=20)
        for _, row in self.df_to_analyze.iterrows():
            executor.submit(self._analyze_keyword_concurrently, row)
        
        # キューのポーリングを開始
        self.dialog.after(100, self._process_queue)
        
        # Tkinterのメインループを開始
        self.dialog.mainloop()
        
        # メインループ終了後、後片付けと最終処理
        executor.shutdown(wait=False)
        self.dialog.destroy()
        self._finalize_and_save()
